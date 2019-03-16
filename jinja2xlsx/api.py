import re
from dataclasses import asdict
from typing import Union, Optional, Dict, Iterator

from openpyxl import Workbook
from openpyxl.cell import MergedCell, Cell
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.worksheet import Worksheet
from requests_html import HTML, Element

from jinja2xlsx.models import Style


def render(html_str: str) -> Workbook:
    html = HTML(html=html_str)

    table = html.find("table", first=True)

    # colgroup = table.find("colgroup", first=True)
    # assert colgroup, "No colgroup with col defined"
    # columns = colgroup.find("col")
    # assert columns, "No colgroup with col defined"

    wb = Workbook()
    ws: Worksheet = wb.active

    table_body = table.find("tbody", first=True)
    fill_sheet_with_table_data(ws, table_body)

    return wb


def fill_sheet_with_table_data(sheet: Worksheet, table: Element) -> None:
    row_index = 0
    col_index = 0

    for row in table.find("tr"):
        for html_cell in row.find("td"):
            target_cell = sheet.cell(row_index + 1, col_index + 1)
            while True:
                if isinstance(target_cell, MergedCell):
                    col_index += 1
                    target_cell = sheet.cell(row_index + 1, col_index + 1)
                else:
                    break

            target_cell.value = parse_cell_value(html_cell.text)

            colspan = int(html_cell.attrs.get("colspan", 1))
            rowspan = int(html_cell.attrs.get("rowspan", 1))
            style = extract_style(html_cell.attrs.get("style"))

            if colspan > 1 or rowspan > 1:
                cell_range = dict(
                    start_row=row_index + 1,
                    start_column=col_index + 1,
                    end_row=row_index + rowspan,
                    end_column=col_index + colspan,
                )
                sheet.merge_cells(**cell_range)
                # style_multiple_cells(cell_range, style)
            else:
                style_single_cell(target_cell, style)

            col_index += colspan

        row_index += 1
        col_index = 0


def parse_cell_value(cell_text: str) -> Union[int, float, str]:
    """
    >>> parse_cell_value("")
    ''
    >>> parse_cell_value("ass")
    'ass'
    >>> parse_cell_value("1")
    1
    >>> parse_cell_value("1.2")
    1.2
    """
    try:
        return int(cell_text)
    except ValueError:
        try:
            return float(cell_text)
        except ValueError:
            return cell_text


def extract_style(style_attr: str) -> Optional[Style]:
    """
    >>> style = extract_style("border: 1px solid black; text-align: center; font-weight: bold")
    >>> style.alignment.horizontal
    'center'
    >>> style.border.left.style
    'thin'
    >>> style.border.left.style == style.border.right.style == style.border.top.style == style.border.bottom.style
    True
    >>> style.font.bold
    True
    """
    if not style_attr:
        return None

    style_dict = {
        style.strip(): value.strip()
        for style, value in (style.split(":") for style in filter(None, style_attr.split(";")))
    }

    border = _build_border(style_dict) or Border()

    text_align_rule = style_dict.get("text-align")
    if text_align_rule:
        alignment = Alignment(horizontal=text_align_rule)
    else:
        alignment = Alignment()

    font = Font()
    if style_dict.get("font-weight") == "bold":
        font.bold = True

    return Style(border, alignment, font)


def _build_border(style_dict: Dict[str, str]) -> Border:
    """
    >>> border = _build_border({"border": "1px solid black"})
    >>> border.left.style
    'thin'
    >>> border.left.style == border.right.style == border.top.style == border.bottom.style
    True
    >>> border = _build_border({"border-right": "2px solid black"})
    >>> border.right.style
    'medium'
    """

    def _from_border_attr(border_attr: str) -> Optional[Border]:
        border_rule = style_dict.get(border_attr)
        if not border_rule:
            return None

        if border_rule == "1px solid black":
            side = Side(style="thin")
        elif re.match(r"\d+px solid black", border_rule):
            side = Side(style="medium")
        else:
            side = Side()

        if border_attr == "border":
            return Border(left=side, right=side, top=side, bottom=side)
        if border_attr == "border-left":
            return Border(left=side)
        if border_attr == "border-right":
            return Border(right=side)
        if border_attr == "border-top":
            return Border(top=side)
        if border_attr == "border-bottom":
            return Border(bottom=side)

        return None

    borders: Iterator[Border] = filter(
        None,
        (
            _from_border_attr("border"),
            _from_border_attr("border-left"),
            _from_border_attr("border-right"),
            _from_border_attr("border-top"),
            _from_border_attr("border-bottom"),
        ),
    )

    return next(borders, None)


def style_single_cell(cell: Cell, style: Optional[Style]) -> None:
    if not style:
        return

    for style_key, value in asdict(style).items():
        setattr(cell, style_key, value)
